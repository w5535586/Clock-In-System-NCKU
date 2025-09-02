import sys
import os
import configparser
from datetime import datetime
from PyQt5.QtWidgets import (
    QApplication, QWidget, QVBoxLayout, QPushButton,
    QLabel, QComboBox, QLineEdit, QMessageBox, QTabWidget,
    QListWidget, QInputDialog, QTableWidget, QTableWidgetItem, QDateTimeEdit, QFileDialog, QDateEdit, QHeaderView
)
from PyQt5.QtCore import QDateTime, QDate, Qt
from PyQt5.QtGui import QPixmap, QFont, QIcon
from openpyxl import Workbook, load_workbook

def get_base_path():
    """取得程式的基準路徑（支援 exe 和開發模式）"""
    if getattr(sys, 'frozen', False):  
        # exe 模式
        return os.path.dirname(sys.executable)
    else:
        # 開發模式 (python script)
        return os.path.dirname(os.path.abspath(__file__))

BASE_PATH = get_base_path()
LAST_PATH_FILE = os.path.join(BASE_PATH, "last_path.txt")

def get_last_config_path():
    if os.path.exists(LAST_PATH_FILE):
        with open(LAST_PATH_FILE, "r", encoding="utf-8") as f:
            return f.read().strip()
    return None

def save_last_config_path(path):
    with open(LAST_PATH_FILE, "w", encoding="utf-8") as f:
        f.write(path)

def get_config_path():
    # 1. 讀取上次使用的路徑
    last_path = get_last_config_path()
    if last_path and os.path.exists(last_path):
        return last_path

    # 2. 固定預設路徑（桌面）
    default_path = os.path.join(os.path.expanduser("~"), "Desktop", "click_in", "config.ini")
    if os.path.exists(default_path):
        save_last_config_path(default_path)
        return default_path

    # 3. 嘗試 exe/py 同目錄下的 config.ini
    local_path = os.path.join(BASE_PATH, "config.ini")
    if os.path.exists(local_path):
        save_last_config_path(local_path)
        return local_path

    # 4. 如果還是沒有 → 回傳 None
    return None



CONFIG_FILE = get_config_path()


class AttendanceSystem(QWidget):
    def __init__(self):
        super().__init__()
        
        global CONFIG_FILE
        if CONFIG_FILE is None:
            # 現在 QApplication 已經建立了，可以用 QFileDialog / QMessageBox
            path, _ = QFileDialog.getOpenFileName(self, "選擇設定檔", "", "INI Files (*.ini)")
            if not path:
                QMessageBox.critical(self, "錯誤", "找不到 config.ini，程式無法繼續執行！")
                sys.exit(1)
            CONFIG_FILE = path
            save_last_config_path(CONFIG_FILE)
            
        self.setWindowTitle("國立成功大學統計系打卡系統")
        self.setWindowIcon(QIcon("ncku_logo.jpg"))
        self.setGeometry(200, 200, 700, 800)

        # 讀取設定（含管理員密碼與目前學期資料夾）
        self.admin_password, self.semester_folder = self.load_config()
        self.staff_file = os.path.join(self.semester_folder, "staff.xlsx")
        self.expected_file = os.path.join(self.semester_folder, "expected.xlsx")
        self.ensure_semester_basics()

        # --- UI 美化 ---
        self.setStyleSheet("""
            QWidget {
                font-family: Microsoft JhengHei;
                font-size: 14px;
            }
            QPushButton {
                background-color: #9B1C1C;
                color: white;
                border-radius: 8px;
                padding: 8px;
            }
            QPushButton:hover {
                background-color: #B22222;
            }
            QTabWidget::pane {
                border: 1px solid #9B1C1C;
                border-radius: 6px;
            }
            QTabBar::tab {
                background: #E5E5E5;
                padding: 8px 16px;
                border-radius: 6px;
                margin: 2px;
            }
            QTabBar::tab:selected {
                background: #9B1C1C;
                color: white;
            }
        """)

        header_layout = QVBoxLayout()

        # --- 分頁 ---
        self.tabs = QTabWidget()
        self.attendance_tab = QWidget()
        self.staff_tab = QWidget()
        self.worktime_tab = QWidget()
        self.duty_tab = QWidget()
        self.tabs.addTab(self.attendance_tab, "打卡系統")
        self.tabs.addTab(self.staff_tab, "人員設定")
        self.tabs.addTab(self.worktime_tab, "工時統計")
        self.tabs.addTab(self.duty_tab, "值班查詢")

        self.init_attendance_tab()
        self.init_staff_tab()
        self.init_worktime_tab()

        # --- 主版面 ---
        layout = QVBoxLayout()
        layout.addLayout(header_layout)
        layout.addWidget(self.tabs)

        self.semester_btn = QPushButton(f"目前學期：{self.semester_folder.split('/')[-1]}")
        self.semester_btn.clicked.connect(self.change_semester)
        layout.addWidget(self.semester_btn)

        self.setLayout(layout)

        # 載入人員資料
        self.load_staff()
        self.load_attendance_records()
        self.init_duty_query_tab()
        self.tabs.setCurrentIndex(0)
        self.tabs.currentChanged.connect(self.check_password)


    # ---------------- 設定檔 ----------------
    def load_config(self):
        if not os.path.exists(CONFIG_FILE):
            QMessageBox.critical(self, "錯誤", f"找不到設定檔 {CONFIG_FILE}，請確認後再試！")
            sys.exit(1)  # 強制結束程式
        
        config = configparser.ConfigParser()
        config.read(CONFIG_FILE, encoding="utf-8")

        pwd = config.get("admin", "password", fallback="1234")
        semester = config.get("system", "semester_folder", fallback="114年上學期")
        os.makedirs(semester, exist_ok=True)
        return pwd, semester


    def save_config(self):
        config = configparser.ConfigParser()
        config["admin"] = {"password": self.admin_password}
        config["system"] = {"semester_folder": self.semester_folder}
        with open(CONFIG_FILE, "w", encoding="utf-8") as f:
            config.write(f)
        save_last_config_path(CONFIG_FILE)  # 這行很重要！


    def ensure_semester_basics(self):
        os.makedirs(self.semester_folder, exist_ok=True)
        # 建立 staff.xlsx
        if not os.path.exists(self.staff_file):
            wb = Workbook()
            ws = wb.active
            ws.append(["姓名"])  # 標題
            wb.save(self.staff_file)
            wb.close()
        # 建立 expected.xlsx
        if not os.path.exists(self.expected_file):
            wb = Workbook()
            ws = wb.active
            ws.append(["姓名", "應到工時"])  # 標題
            wb.save(self.expected_file)
            wb.close()

    def change_semester(self):
        folder = QFileDialog.getExistingDirectory(self, "選擇學期資料夾")
        if folder:
            self.semester_folder = folder
            self.semester_btn.setText(f"目前學期：{self.semester_folder.split('/')[-1]}")
            self.staff_file = os.path.join(self.semester_folder, "staff.xlsx")
            self.expected_file = os.path.join(self.semester_folder, "expected.xlsx")
            self.ensure_semester_basics()
            self.save_config()
            # 重新載入該學期的資料
            self.load_staff()
            self.load_expected_worktime()
            self.refresh_worktime_name_combo()
            self.load_attendance_records()
            self.refresh_duty_name_combo()

    # ---------------- 驗證密碼 ----------------
    def check_password(self, index):
        if self.tabs.tabText(index) == "人員設定":
            pwd, ok = QInputDialog.getText(self, "密碼驗證", "請輸入管理員密碼：", QLineEdit.Password)
            if not ok or pwd != self.admin_password:
                QMessageBox.warning(self, "錯誤", "密碼錯誤，無法進入人員設定！")
                self.tabs.setCurrentIndex(0)

    # ---------------- 人員設定 ----------------
    def init_staff_tab(self):
        layout = QVBoxLayout()
        self.staff_list = QListWidget()
        layout.addWidget(QLabel("本學期人員："))
        layout.addWidget(self.staff_list)

        self.name_input = QLineEdit()
        self.name_input.setPlaceholderText("輸入人員姓名")
        layout.addWidget(self.name_input)

        add_btn = QPushButton("新增人員")
        add_btn.clicked.connect(self.add_staff)
        layout.addWidget(add_btn)

        del_btn = QPushButton("刪除選擇人員")
        del_btn.clicked.connect(self.delete_staff)
        layout.addWidget(del_btn)

        self.staff_tab.setLayout(layout)

    def load_staff(self):
        self.staff_list.clear()
        self.staff_combo.clear()
        if os.path.exists(self.staff_file):
            wb = load_workbook(self.staff_file)
            ws = wb.active
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row and row[0]:
                    self.staff_list.addItem(row[0])
                    self.staff_combo.addItem(row[0])
            wb.close()

    def add_staff(self):
        name = self.name_input.text().strip()
        if not name:
            QMessageBox.warning(self, "錯誤", "姓名不能為空！")
            return
        # 確保名單檔存在
        if not os.path.exists(self.staff_file):
            wb = Workbook()
            ws = wb.active
            ws.append(["姓名"])  # 標題
            wb.save(self.staff_file)
        # 檢查是否重複
        wb = load_workbook(self.staff_file)
        ws = wb.active
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row and row[0] == name:
                QMessageBox.warning(self, "錯誤", "此人員已存在！")
                wb.close()
                return
        ws.append([name])
        wb.save(self.staff_file)
        wb.close()
        # UI 更新
        self.staff_list.addItem(name)
        self.staff_combo.addItem(name)
        self.worktime_name_combo.addItem(name)
        self.query_name_combo.addItem(name)
        self.name_input.clear()

    def delete_staff(self):
        selected = self.staff_list.currentItem()
        if not selected:
            QMessageBox.warning(self, "錯誤", "請選擇要刪除的人員！")
            return
        name = selected.text()
        if not os.path.exists(self.staff_file):
            return
        wb = load_workbook(self.staff_file)
        ws = wb.active
        for row in ws.iter_rows(min_row=2):
            if row[0].value == name:
                ws.delete_rows(row[0].row)
                break
        wb.save(self.staff_file)
        wb.close()
        # UI 更新
        self.staff_list.takeItem(self.staff_list.row(selected))
        idx = self.staff_combo.findText(name)
        if idx >= 0:
            self.staff_combo.removeItem(idx)
        idx2 = self.worktime_name_combo.findText(name)
        if idx2 >= 0:
            self.worktime_name_combo.removeItem(idx2)
        idx3 = self.query_name_combo.findText(name)   # 🔹 新增這段
        if idx3 >= 0:
            self.query_name_combo.removeItem(idx3)

    # ---------------- 打卡系統 ----------------
    def init_attendance_tab(self):
        layout = QVBoxLayout()
        layout.addWidget(QLabel("選擇打卡人員："))
        self.staff_combo = QComboBox()
        layout.addWidget(self.staff_combo)

        # 即時打卡
        sign_in_btn = QPushButton("簽到")
        sign_in_btn.clicked.connect(lambda: self.record_attendance("簽到", is_manual=False))
        layout.addWidget(sign_in_btn)
        sign_out_btn = QPushButton("簽退")
        sign_out_btn.clicked.connect(lambda: self.record_attendance("簽退", is_manual=False))
        layout.addWidget(sign_out_btn)

        # 補打卡
        layout.addWidget(QLabel("補打卡（需要管理員密碼）："))
        self.datetime_edit = QDateTimeEdit(QDateTime.currentDateTime())
        self.datetime_edit.setDisplayFormat("yyyy-MM-dd HH:mm:ss")
        self.datetime_edit.setCalendarPopup(True)
        self.datetime_edit.setMaximumDateTime(QDateTime.currentDateTime())
        layout.addWidget(self.datetime_edit)
        manual_sign_in_btn = QPushButton("補簽到")
        manual_sign_in_btn.clicked.connect(lambda: self.record_attendance("簽到", is_manual=True))
        layout.addWidget(manual_sign_in_btn)
        manual_sign_out_btn = QPushButton("補簽退")
        manual_sign_out_btn.clicked.connect(lambda: self.record_attendance("簽退", is_manual=True))
        layout.addWidget(manual_sign_out_btn)

        # 打卡紀錄表
        layout.addWidget(QLabel("最新 10 筆打卡紀錄："))
        self.record_table = QTableWidget()
        self.record_table.setColumnCount(4)
        self.record_table.setHorizontalHeaderLabels(["姓名", "動作", "日期", "時間"])
        header = self.record_table.horizontalHeader()
        header.setSectionResizeMode(0, QHeaderView.ResizeToContents)  # 姓名自動調整
        header.setSectionResizeMode(1, QHeaderView.ResizeToContents)  # 動作自動調整
        header.setSectionResizeMode(2, QHeaderView.ResizeToContents)  # 日期自動調整
        header.setSectionResizeMode(3, QHeaderView.Stretch)           # 時間填滿剩下空間
        layout.addWidget(self.record_table)
        self.staff_combo.currentIndexChanged.connect(self.load_attendance_records)
        del_btn = QPushButton("刪除選擇紀錄")
        del_btn.clicked.connect(self.delete_selected_record)
        layout.addWidget(del_btn)

        
        self.attendance_tab.setLayout(layout)

    def get_attendance_file(self, name):
        return os.path.join(self.semester_folder, f"{name}.xlsx")

    def load_attendance_records(self):
        """顯示目前選擇人員的打卡紀錄（依 Excel 順序）。"""
        self.record_table.setRowCount(0)
        records = []
        name = self.staff_combo.currentText()
        if not name:
            return
        file = self.get_attendance_file(name)
        if os.path.exists(file):
            w = load_workbook(file)
            s = w.active
            for r in s.iter_rows(min_row=2, values_only=True):
                records.append(r)  # [姓名, 動作, 日期, 時間]
            w.close()

        # 顯示（依原始順序，不排序、不倒序）
        records = records[-10:]
        self.record_table.setRowCount(len(records))
        for i, row in enumerate(records):
            for j, value in enumerate(row):
                item = QTableWidgetItem(str(value))
                item.setTextAlignment(Qt.AlignCenter)  # 新增這行
                self.record_table.setItem(i, j, item)
        self.record_table.resizeRowsToContents()
        # self.record_table.resizeColumnsToContents()

        
    def delete_selected_record(self):
        row = self.record_table.currentRow()
        if row < 0:
            QMessageBox.warning(self, "錯誤", "請先選擇要刪除的紀錄！")
            return

        name_item = self.record_table.item(row, 0)
        action_item = self.record_table.item(row, 1)
        date_item = self.record_table.item(row, 2)
        time_item = self.record_table.item(row, 3)

        if not (name_item and action_item and date_item and time_item):
            QMessageBox.warning(self, "錯誤", "選擇的資料不完整！")
            return

        name = name_item.text()
        action = action_item.text()
        date = date_item.text()
        timestamp = time_item.text()

        file = self.get_attendance_file(name)
        if not os.path.exists(file):
            QMessageBox.warning(self, "錯誤", "找不到該人員的檔案！")
            return

        # 確認刪除
        reply = QMessageBox.question(
            self, "確認刪除", f"確定要刪除 {name} {action} {timestamp} 這筆紀錄嗎？",
            QMessageBox.Yes | QMessageBox.No
        )
        if reply != QMessageBox.Yes:
            return

        # 開檔並刪除對應列
        wb = load_workbook(file)
        ws = wb.active
        for r in ws.iter_rows(min_row=2):
            if (r[0].value == name and r[1].value == action 
                    and r[2].value == date and r[3].value == timestamp):
                ws.delete_rows(r[0].row, 1)
                break
        wb.save(file)
        wb.close()

        QMessageBox.information(self, "成功", "紀錄已刪除！")
        self.load_attendance_records()
        self.record_table.resizeRowsToContents()




    def record_attendance(self, action, is_manual=False):
        name = self.staff_combo.currentText()
        if not name:
            QMessageBox.warning(self, "錯誤", "請先設定人員！")
            return
        # 補打卡需密碼
        if is_manual:
            pwd, ok = QInputDialog.getText(self, "密碼驗證", "請輸入管理員密碼：", QLineEdit.Password)
            if not ok or pwd != self.admin_password:
                QMessageBox.warning(self, "錯誤", "密碼錯誤，無法執行補打卡！")
                return
        # 時間
        if is_manual:
            dt = self.datetime_edit.dateTime().toPyDateTime()
            now_date = dt.strftime("%Y-%m-%d")
            now_time = dt.strftime("%Y-%m-%d %H:%M:%S")
        else:
            dt = datetime.now()
            now_date = dt.strftime("%Y-%m-%d")
            now_time = dt.strftime("%Y-%m-%d %H:%M:%S")

        # 個人檔案
        file = self.get_attendance_file(name)
        if not os.path.exists(file):
            wb = Workbook()
            ws = wb.active
            ws.append(["姓名", "動作", "日期", "時間"])
            wb.save(file)
            wb.close()

        wb = load_workbook(file)
        ws = wb.active

        last_action = None
        last_signin_time = None
        for row in ws.iter_rows(min_row=2, values_only=True):
            last_action = row[1]
            if row[1] == "簽到":
                last_signin_time = datetime.strptime(row[3], "%Y-%m-%d %H:%M:%S")
            elif row[1] == "簽退":
                last_signin_time = None  # 一組簽到退結束

        # 防呆檢查
        if action == "簽到" and last_action == "簽到":
            QMessageBox.warning(self, "錯誤", f"{name} 上次已簽到，必須先簽退才能再簽到！")
            wb.close()
            return
        if action == "簽退":
            if last_action is None or last_action == "簽退":
                QMessageBox.warning(self, "錯誤", f"{name} 尚未簽到，或已簽退過，不能直接簽退！")
                wb.close()
                return
            if last_signin_time and dt <= last_signin_time:
                QMessageBox.warning(self, "錯誤", f"{name} 的簽退時間必須晚於簽到時間！")
                wb.close()
                return

        # 寫入
        ws.append([name, action, now_date, now_time])
        wb.save(file)
        wb.close()
        QMessageBox.information(self, "成功", f"{name} 已完成 {action}！")
        self.load_attendance_records()


    # ---------------- 工時 ----------------
    def init_worktime_tab(self):
        layout = QVBoxLayout()
        layout.addWidget(QLabel("工時統計："))
        self.worktime_table = QTableWidget()
        self.worktime_table.setColumnCount(4)
        self.worktime_table.setHorizontalHeaderLabels(["姓名", "應到工時(小時)", "實際工時(小時)", "差異(小時)"])
        header = self.worktime_table.horizontalHeader()
        header.setSectionResizeMode(QHeaderView.ResizeToContents)
        header.setStretchLastSection(True)
        layout.addWidget(self.worktime_table)

        layout.addWidget(QLabel("設定人員應到工時："))
        self.worktime_name_combo = QComboBox()
        layout.addWidget(self.worktime_name_combo)
        self.worktime_expected_input = QLineEdit()
        self.worktime_expected_input.setPlaceholderText("輸入應到工時（小時）")
        layout.addWidget(self.worktime_expected_input)
        save_btn = QPushButton("儲存應到工時")
        save_btn.clicked.connect(self.save_expected_worktime)
        layout.addWidget(save_btn)
        calc_btn = QPushButton("重新計算工時")
        calc_btn.clicked.connect(self.calculate_worktime)
        layout.addWidget(calc_btn)
        export_btn = QPushButton("匯出結果到 Excel")
        export_btn.clicked.connect(lambda: self.calculate_worktime(export=True))
        layout.addWidget(export_btn)

        self.worktime_tab.setLayout(layout)
        # 準備資料
        self.expected_worktime = {}
        self.load_expected_worktime()
        self.refresh_worktime_name_combo()
        
        layout.addWidget(QLabel("選擇匯出起始日期（可留空）："))
        self.export_start_date = QDateEdit()
        self.export_start_date.setDisplayFormat("yyyy-MM-dd")
        self.export_start_date.setCalendarPopup(True)
        self.export_start_date.setDate(QDate(2000, 1, 1))  # 設一個初始值
        self.export_start_date.clear()  # 預設清空
        layout.addWidget(self.export_start_date)

        layout.addWidget(QLabel("選擇匯出結束日期（可留空）："))
        self.export_end_date = QDateEdit()
        self.export_end_date.setDisplayFormat("yyyy-MM-dd")
        self.export_end_date.setCalendarPopup(True)
        self.export_end_date.setDate(QDate.currentDate())
        self.export_end_date.clear()  # 預設清空
        layout.addWidget(self.export_end_date)

    def refresh_worktime_name_combo(self):
        self.worktime_name_combo.clear()
        if not os.path.exists(self.staff_file):
            return
        wb = load_workbook(self.staff_file)
        ws = wb.active
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row and row[0]:
                self.worktime_name_combo.addItem(row[0])
        wb.close()

    def calculate_worktime(self, export=False):
        # 讀取日期篩選（允許空白）
        start_date = self.export_start_date.date().toPyDate() if self.export_start_date.date().isValid() else None
        end_date = self.export_end_date.date().toPyDate() if self.export_end_date.date().isValid() else None

        worktime = {}
        if os.path.exists(self.staff_file):
            wb = load_workbook(self.staff_file)
            ws = wb.active
            for row in ws.iter_rows(min_row=2, values_only=True):
                name = row[0]
                if not name:
                    continue
                file = self.get_attendance_file(name)
                if os.path.exists(file):
                    w = load_workbook(file)
                    s = w.active
                    last_signin = None
                    for r in s.iter_rows(min_row=2, values_only=True):
                        _, action, _, timestamp = r
                        dt = datetime.strptime(timestamp, "%Y-%m-%d %H:%M:%S")
                        if action == "簽到":
                            last_signin = dt
                        elif action == "簽退" and last_signin:
                            # 檢查日期篩選
                            if (start_date is None or last_signin.date() >= start_date) and \
                               (end_date is None or dt.date() <= end_date):
                                delta = dt - last_signin
                                minutes = int(delta.total_seconds() // 60)
                                worktime[name] = worktime.get(name, 0) + minutes
                            last_signin = None
                    w.close()
            wb.close()

        # 若無打卡，但 expected 有資料，也要顯示
        for name in self.expected_worktime.keys():
            if name not in worktime:
                worktime[name] = 0
        # 顯示表格與待匯出資料
        self.worktime_table.setRowCount(0)
        results = []
        for i, (name, minutes) in enumerate(worktime.items()):
            expected_hours = self.expected_worktime.get(name, 0.0)
            expected_minutes = int(expected_hours * 60)
            diff_minutes = minutes - expected_minutes
            self.worktime_table.insertRow(i)
            self.worktime_table.setItem(i, 0, QTableWidgetItem(name))
            self.worktime_table.item(i, 0).setTextAlignment(Qt.AlignCenter)

            self.worktime_table.setItem(i, 1, QTableWidgetItem(self.format_minutes(expected_minutes)))
            self.worktime_table.item(i, 1).setTextAlignment(Qt.AlignCenter)

            self.worktime_table.setItem(i, 2, QTableWidgetItem(self.format_minutes(minutes)))
            self.worktime_table.item(i, 2).setTextAlignment(Qt.AlignCenter)

            self.worktime_table.setItem(i, 3, QTableWidgetItem(self.format_minutes(diff_minutes, show_sign=True)))
            self.worktime_table.item(i, 3).setTextAlignment(Qt.AlignCenter)

            results.append([name,
                            self.format_minutes(expected_minutes),
                            self.format_minutes(minutes),
                            self.format_minutes(diff_minutes, show_sign=True)])
        # 匯出
        if export:
            wb = Workbook()
            ws = wb.active
            ws.title = "工時總表"
            ws.append(["姓名", "應到工時", "實際工時", "差異"])
            for row in results:
                ws.append(row)

            # 匯出時段
            start_date = self.export_start_date.date().toPyDate() if self.export_start_date.date().isValid() else None
            end_date = self.export_end_date.date().toPyDate() if self.export_end_date.date().isValid() else None

            # 為每個人建立副表單
            for name in worktime.keys():
                sub_ws = wb.create_sheet(title=name)
                sub_ws.append(["姓名", "簽到時間", "簽退時間", "值班時長"])

                file = self.get_attendance_file(name)
                if os.path.exists(file):
                    pwb = load_workbook(file)
                    ps = pwb.active
                    last_signin = None
                    for r in ps.iter_rows(min_row=2, values_only=True):
                        _, action, date_str, timestamp = r
                        dt = datetime.strptime(timestamp, "%Y-%m-%d %H:%M:%S")
                        if action == "簽到":
                            last_signin = dt
                        elif action == "簽退" and last_signin:
                            if (start_date is None or last_signin.date() >= start_date) and \
                               (end_date is None or dt.date() <= end_date):
                                delta = dt - last_signin
                                hours = int(delta.total_seconds() // 3600)
                                minutes = int((delta.total_seconds() % 3600) // 60)
                                duration = f"{hours:02d}:{minutes:02d}"
                                sub_ws.append([name,
                                               last_signin.strftime("%Y-%m-%d %H:%M:%S"),
                                               dt.strftime("%Y-%m-%d %H:%M:%S"),
                                               duration])
                            last_signin = None
                    pwb.close()

            # 儲存
            default_name = f"{os.path.basename(self.semester_folder)}_worktime_result.xlsx"
            file_path, _ = QFileDialog.getSaveFileName(
                self, "儲存工時計算結果",
                os.path.join(self.semester_folder, default_name),
                "Excel Files (*.xlsx)"
            )
            if file_path:
                wb.save(file_path)
                QMessageBox.information(self, "匯出完成", f"工時計算結果已輸出到 {file_path}")
            wb.close()



    def format_minutes(self, minutes, show_sign=False):
        sign = ""
        if show_sign:
            if minutes > 0:
                sign = "+"
            elif minutes < 0:
                sign = "-"
            minutes = abs(minutes)
        h = minutes // 60
        m = minutes % 60
        return f"{sign}{h:02d}:{m:02d}"

    def load_expected_worktime(self):
        # 確保檔案存在
        if not os.path.exists(self.expected_file):
            self.ensure_semester_basics()
        wb = load_workbook(self.expected_file)
        ws = wb.active
        self.expected_worktime = {}
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row and row[0]:
                try:
                    self.expected_worktime[row[0]] = float(row[1]) if row[1] is not None else 0.0
                except Exception:
                    self.expected_worktime[row[0]] = 0.0
        wb.close()

    def save_expected_worktime(self):
        name = self.worktime_name_combo.currentText().strip()
        try:
            expected = float(self.worktime_expected_input.text())
        except Exception:
            QMessageBox.warning(self, "錯誤", "應到工時必須是數字！")
            return
        if not name:
            QMessageBox.warning(self, "錯誤", "請先選擇姓名！")
            return
        # 更新記憶體
        self.expected_worktime[name] = expected
        # 寫檔
        if not os.path.exists(self.expected_file):
            self.ensure_semester_basics()
        wb = load_workbook(self.expected_file)
        ws = wb.active
        found = False
        for row in ws.iter_rows(min_row=2):
            if row[0].value == name:
                row[1].value = expected
                found = True
                break
        if not found:
            ws.append([name, expected])
        wb.save(self.expected_file)
        wb.close()
        QMessageBox.information(self, "成功", f"{name} 的應到工時已設定為 {expected} 小時！")
        self.calculate_worktime()

    def init_duty_query_tab(self):
        layout = QVBoxLayout()
        layout.addWidget(QLabel("選擇人員："))

        self.query_name_combo = QComboBox()
        layout.addWidget(self.query_name_combo)

        # 查詢日期範圍
        layout.addWidget(QLabel("選擇查詢起始日期："))
        self.start_date = QDateTimeEdit(QDateTime.currentDateTime())
        self.start_date.setDisplayFormat("yyyy-MM-dd")
        self.start_date.setCalendarPopup(True)
        layout.addWidget(self.start_date)

        layout.addWidget(QLabel("選擇查詢結束日期："))
        self.end_date = QDateTimeEdit(QDateTime.currentDateTime())
        self.end_date.setDisplayFormat("yyyy-MM-dd")
        self.end_date.setCalendarPopup(True)
        layout.addWidget(self.end_date)

        # 查詢按鈕
        query_btn = QPushButton("查詢值班紀錄")
        query_btn.clicked.connect(self.load_duty_records)
        layout.addWidget(query_btn)

        # 顯示表格
        self.duty_table = QTableWidget()
        self.duty_table.setColumnCount(4)
        self.duty_table.setHorizontalHeaderLabels(["姓名", "簽到時間", "簽退時間", "值班時長"])
        header = self.duty_table.horizontalHeader()
        header.setSectionResizeMode(QHeaderView.ResizeToContents)
        header.setStretchLastSection(True)
        layout.addWidget(self.duty_table)

        self.duty_tab.setLayout(layout)

        # 載入人員
        self.refresh_duty_name_combo()
    
    def refresh_duty_name_combo(self):
        self.query_name_combo.clear()
        if os.path.exists(self.staff_file):
            wb = load_workbook(self.staff_file)
            ws = wb.active
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row and row[0]:
                    self.query_name_combo.addItem(row[0])
            wb.close()

    def load_duty_records(self):
        """查詢某人的所有簽到簽退組合，依照簽到日期排序，可篩選日期區間"""
        name = self.query_name_combo.currentText()
        if not name:
            return

        file = self.get_attendance_file(name)
        if not os.path.exists(file):
            QMessageBox.warning(self, "錯誤", f"找不到 {name} 的打卡資料！")
            return

        start_date = self.start_date.date().toPyDate()
        end_date = self.end_date.date().toPyDate()

        # 防呆：起始時間必須早於結束時間
        if start_date > end_date:
            QMessageBox.warning(self, "錯誤", "起始日期不能晚於結束日期！")
            return

        wb = load_workbook(file)
        ws = wb.active

        records = []
        last_signin = None
        for row in ws.iter_rows(min_row=2, values_only=True):
            _, action, date_str, timestamp = row
            dt = datetime.strptime(timestamp, "%Y-%m-%d %H:%M:%S")
            if action == "簽到":
                last_signin = dt
            elif action == "簽退" and last_signin:
                # 比對日期（只看年月日）
                if start_date <= last_signin.date() <= end_date or start_date <= dt.date() <= end_date:
                    delta = dt - last_signin
                    hours = delta.total_seconds() // 3600
                    minutes = (delta.total_seconds() % 3600) // 60
                    duration = f"{int(hours):02d}:{int(minutes):02d}"
                    records.append([name,
                                    last_signin.strftime("%Y-%m-%d %H:%M:%S"),
                                    dt.strftime("%Y-%m-%d %H:%M:%S"),
                                    duration])
                last_signin = None
        wb.close()

        # 依照簽到時間排序
        records.sort(key=lambda x: datetime.strptime(x[1], "%Y-%m-%d %H:%M:%S"))

        # 顯示到表格
        self.duty_table.setRowCount(len(records))
        for i, row in enumerate(records):
            for j, value in enumerate(row):
                item = QTableWidgetItem(str(value))
                item.setTextAlignment(Qt.AlignCenter)
                self.duty_table.setItem(i, j, item)
        self.duty_table.resizeColumnsToContents()



if __name__ == "__main__":
    app = QApplication(sys.argv)
    window = AttendanceSystem()
    window.show()
    sys.exit(app.exec_())
